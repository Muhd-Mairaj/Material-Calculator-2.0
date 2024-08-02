import copy

from openpyxl.worksheet import _read_only
from openpyxl.worksheet.worksheet import Worksheet

from helper import deprecated


class MySheet():
    def _my_setup(self):
        self.header = self[self.min_row]

        self._profile_column = 0
        self._length_column = 0
        self._qty_column = 0
        # start at 1 because indexing is 1-based
        for i, cell in enumerate(self.header, 1):
            cell_value = cell.value
            if cell_value == None:
                continue

            elif cell_value.lower() == "profile":
                self._profile_column = i
            elif cell_value.lower() == "length":
                self._length_column = i
            elif cell_value.lower() == "qty.":
                self._qty_column = i

        self._profiles: dict[str, dict[int, int]] = {}
        profile_iterator = self.iter_rows(min_row=2, min_col=self._profile_column, max_col=self._profile_column)
        length_iterator = self.iter_rows(min_row=2, min_col=self._length_column, max_col=self._length_column)
        qty_iterator = self.iter_rows(min_row=2, min_col=self._qty_column, max_col=self._qty_column)
        for profile_, length_, qty_ in zip(profile_iterator, length_iterator, qty_iterator):
            profile_value = profile_[0].value
            length_value = length_[0].value
            qty_value = qty_[0].value

            self._profiles.setdefault(profile_value, {})
            self._profiles[profile_value][length_value] = int(qty_value)

    def get_items(self, profile: str) -> dict[int, int]:
        return self._profiles[profile].copy()

    @property
    def profiles(self) -> tuple[str]:
        return tuple(self._profiles.keys())


class MyWorksheet(Worksheet, MySheet):
    # def __init__(self, parent: _read_only.Workbook | None, title: str | _write_only._Decodable | None = None) -> None:
    #     super().__init__(parent, title)
    def __init__(self, sheet: Worksheet) -> None:
        raise NotImplementedError


class MyReadOnlyWorksheet(_read_only.ReadOnlyWorksheet, MySheet):
    # def __init__(self, parent_workbook: _read_only.Workbook, title: str, worksheet_path, shared_strings: _read_only.SupportsGetItem[int, str]) -> None:
    #     super().__init__(parent_workbook, title, worksheet_path, shared_strings)
    def __init__(self, read_only_worksheet: _read_only.ReadOnlyWorksheet) -> None:
        self.parent = read_only_worksheet.parent
        self.title = read_only_worksheet.title
        self.sheet_state = read_only_worksheet.sheet_state
        self._current_row = read_only_worksheet._current_row
        self._worksheet_path = read_only_worksheet._worksheet_path
        self._shared_strings = read_only_worksheet._shared_strings
        self._get_size = read_only_worksheet._get_size
        self.defined_names = read_only_worksheet.defined_names

        self._my_setup()


def get_sheet(sheet: Worksheet | _read_only.ReadOnlyWorksheet) -> MyWorksheet | MyReadOnlyWorksheet:
    """This is the method that should be used to encapsulate the openpyxl worksheet with the required class.


    Args:
        sheet (Worksheet | ReadOnlyWorksheet): The worksheet that needs to be encapsulated

    Returns:
        MyWorksheet | MyReadOnlyWorksheet: The encapsulated worksheet with additional properties
    """
    if isinstance(sheet, Worksheet):
        return MyWorksheet(sheet)

    if isinstance(sheet, _read_only.ReadOnlyWorksheet):
        return MyReadOnlyWorksheet(sheet)


@deprecated
class SheetAnalyser():
    def __init__(self, ws: Worksheet) -> None:
        self._ws: Worksheet = ws
        self._setup()

    def _setup(self):
        self.header = self.ws[self.ws.min_row]

        self._profile_column = 0
        self._length_column = 0
        self._qty_column = 0
        # start at 1 because indexing is 1-based
        for i, cell in enumerate(self.header, 1):
            cell_value = cell.value
            if cell_value == None:
                continue

            elif cell_value.lower() == "profile":
                self._profile_column = i
            elif cell_value.lower() == "length":
                self._length_column = i
            elif cell_value.lower() == "qty.":
                self._qty_column = i

        self._profiles: dict[str, dict[int, int]] = {}
        profile_iterator = self.ws.iter_rows(min_row=2, min_col=self._profile_column, max_col=self._profile_column)
        length_iterator = self.ws.iter_rows(min_row=2, min_col=self._length_column, max_col=self._length_column)
        qty_iterator = self.ws.iter_rows(min_row=2, min_col=self._qty_column, max_col=self._qty_column)
        for profile_, length_, qty_ in zip(profile_iterator, length_iterator, qty_iterator):
            profile_value = profile_[0].value
            length_value = length_[0].value
            qty_value = qty_[0].value

            self._profiles.setdefault(profile_value, {})
            self._profiles[profile_value][length_value] = int(qty_value)

    def get_items(self, profile: str) -> dict[int, int]:
        return self._profiles[profile].copy()

    # def get_items(self, profile: str) -> dict[int, int]:
    #     profile_iterator = self.ws.iter_rows(min_row=2, min_col=self._profile_column, max_col=self._profile_column)
    #     length_iterator = self.ws.iter_rows(min_row=2, min_col=self._length_column, max_col=self._length_column)
    #     qty_iterator = self.ws.iter_rows(min_row=2, min_col=self._qty_column, max_col=self._qty_column)

    #     items = {}
    #     for profile_, length_, qty_ in zip(profile_iterator, length_iterator, qty_iterator):
    #         profile_value = profile_[0].value
    #         length_value = length_[0].value
    #         qty_value = qty_[0].value

    #         if profile_value == profile:
    #             items[length_value] = int(qty_value)

    #     return items

    @property
    def ws(self) -> Worksheet:
        return self._ws

    # @property
    # def profiles(self):
    #     return self._profiles.keys()

    @property
    def profiles(self) -> dict[str, dict[int, int]]:
        return copy.deepcopy(self._profiles)
