import copy

from openpyxl.worksheet.worksheet import Worksheet


class SheetAnalyser():
    def __init__(self, ws: Worksheet) -> None:
        self._ws: Worksheet = ws
        self._setup()

    def _setup(self):
        self.header = self.ws[self.ws.min_row]

        self._profile_column = 0
        self._length_column = 0
        self._qty_column = 0
        # start at 1 because indexing is 1-based
        for i, cell in enumerate(self.header, 1):
            cell_value = cell.value
            if cell_value == None:
                continue

            elif cell_value.lower() == "profile":
                self._profile_column = i
            elif cell_value.lower() == "length":
                self._length_column = i
            elif cell_value.lower() == "qty.":
                self._qty_column = i

        self._profiles: dict[str, dict[int, int]] = {}
        profile_iterator = self.ws.iter_rows(min_row=2, min_col=self._profile_column, max_col=self._profile_column)
        length_iterator = self.ws.iter_rows(min_row=2, min_col=self._length_column, max_col=self._length_column)
        qty_iterator = self.ws.iter_rows(min_row=2, min_col=self._qty_column, max_col=self._qty_column)
        for profile_, length_, qty_ in zip(profile_iterator, length_iterator, qty_iterator):
            profile_value = profile_[0].value
            length_value = length_[0].value
            qty_value = qty_[0].value

            self._profiles.setdefault(profile_value, {})
            self._profiles[profile_value][length_value] = int(qty_value)

    def get_items(self, profile: str) -> dict[int, int]:
        return self._profiles[profile].copy()

    # def get_items(self, profile: str) -> dict[int, int]:
    #     profile_iterator = self.ws.iter_rows(min_row=2, min_col=self._profile_column, max_col=self._profile_column)
    #     length_iterator = self.ws.iter_rows(min_row=2, min_col=self._length_column, max_col=self._length_column)
    #     qty_iterator = self.ws.iter_rows(min_row=2, min_col=self._qty_column, max_col=self._qty_column)

    #     items = {}
    #     for profile_, length_, qty_ in zip(profile_iterator, length_iterator, qty_iterator):
    #         profile_value = profile_[0].value
    #         length_value = length_[0].value
    #         qty_value = qty_[0].value

    #         if profile_value == profile:
    #             items[length_value] = int(qty_value)

    #     return items

    @property
    def ws(self) -> Worksheet:
        return self._ws

    # @property
    # def profiles(self):
    #     return self._profiles.keys()

    @property
    def profiles(self) -> dict[str, dict[int, int]]:
        return copy.deepcopy(self._profiles)
